from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from training_matrix_agent.models.schemas import MatrixRow


class TrainingMatrixService:
    def load_rows(self, workbook_path: Path, sheet_name: str = "Matrix") -> list[MatrixRow]:
        wb = load_workbook(workbook_path)
        ws = wb[sheet_name]
        rows: list[MatrixRow] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            staff_id, full_name, training_date = row[:3]
            rows.append(
                MatrixRow(
                    staff_id=str(staff_id),
                    full_name=str(full_name),
                    training_date=training_date if isinstance(training_date, date) else None,
                )
            )
        return rows

    def apply_updates(self, workbook_path: Path, updates: dict[str, date], sheet_name: str = "Matrix") -> None:
        wb = load_workbook(workbook_path)
        ws = wb[sheet_name]
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            staff_id = str(row[0])
            if staff_id in updates:
                ws.cell(row=row_idx, column=3, value=updates[staff_id])
        wb.save(workbook_path)
