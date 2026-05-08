from __future__ import annotations

from pathlib import Path

import openpyxl


def load_workbook(path: str):
    return openpyxl.load_workbook(Path(path))


def get_staff_names(ws) -> list[str]:
    names = []
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=1).value
        if value:
            names.append(str(value))
    return names
